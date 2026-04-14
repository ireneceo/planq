"""
Priority Q&A 업로드 파서 — 다양한 포맷에서 Q&A 쌍 추출.

지원:
- CSV / TSV: 컬럼명 alias 허용 (question/질문/Q, answer/답변/A, short_answer/짧은답변, keywords, category)
- XLSX: openpyxl, 첫 시트만 사용, 컬럼명 alias 동일
- JSON: [{"question": ..., "answer": ...}] 배열 또는 {"qa_pairs": [...]} 형태
- TXT / MD: 정규식 패턴 — Q:/A:, 질문:/답변:, Q1./A1., ### Q: / ### A: 등 + LLM fallback
- PDF / DOCX: services/extractors.py 로 텍스트 추출 후 LLM extract_qa_pairs_from_text

Returns: list of {question, answer, short_answer|None, keywords|None, category|None}
"""
import csv
import io
import json
import logging
import re
from typing import Optional

logger = logging.getLogger('q-note.qa_parser')


# 확장자 → 파서 종류
PARSER_MAP = {
  'csv': 'tabular',
  'tsv': 'tabular',
  'xlsx': 'xlsx',
  'xls': 'xlsx',
  'json': 'json',
  'txt': 'text',
  'md': 'text',
  'pdf': 'document',
  'docx': 'document',
  'doc': 'document',
}


# 컬럼명 alias — 공백/대소문자 무시
COL_ALIASES: dict[str, list[str]] = {
  'question': ['question', 'q', 'questions', '질문', '문제', '물음', 'ques', '질의'],
  'answer': ['answer', 'a', 'answers', '답변', '답', '응답', 'ans', 'response'],
  'short_answer': ['short_answer', 'short', 'shortanswer', '짧은답변', '1문장', '요약답변', 'summary'],
  'keywords': ['keywords', 'keyword', '키워드', 'tags', 'tag'],
  'category': ['category', '카테고리', '분류', '주제', 'topic', 'section'],
}

_ALIAS_LOOKUP: dict[str, str] = {}
for canon, aliases in COL_ALIASES.items():
  for a in aliases:
    _ALIAS_LOOKUP[a.lower().replace(' ', '').replace('_', '')] = canon


def _normalize_col(name: str) -> Optional[str]:
  if not name:
    return None
  key = str(name).strip().lower().replace(' ', '').replace('_', '')
  return _ALIAS_LOOKUP.get(key)


def _decode_text(raw: bytes) -> str:
  for enc in ('utf-8-sig', 'utf-8', 'cp949', 'euc-kr', 'latin-1'):
    try:
      return raw.decode(enc)
    except UnicodeDecodeError:
      continue
  raise ValueError('텍스트 디코딩 실패 (지원 인코딩: utf-8, cp949, euc-kr)')


def _row_to_qa(row: dict) -> Optional[dict]:
  """dict row (이미 canonical 컬럼으로 매핑된) → qa dict. 필수 필드 체크."""
  q = (row.get('question') or '').strip() if row.get('question') else ''
  a = (row.get('answer') or '').strip() if row.get('answer') else ''
  if not q or not a:
    return None
  if len(q) > 2000 or len(a) > 5000:
    return None
  sa = row.get('short_answer')
  kw = row.get('keywords')
  cat = row.get('category')
  return {
    'question': q,
    'answer': a,
    'short_answer': (sa.strip()[:500] or None) if isinstance(sa, str) else None,
    'keywords': (kw.strip()[:500] or None) if isinstance(kw, str) else None,
    'category': (cat.strip()[:100] or None) if isinstance(cat, str) else None,
  }


# ─────────────────────────────────────────────────────────
# CSV / TSV
# ─────────────────────────────────────────────────────────

def parse_tabular(raw: bytes, filename: str) -> tuple[list[dict], list[str]]:
  text = _decode_text(raw)
  # sniff delimiter (csv vs tsv)
  delimiter = '\t' if filename.lower().endswith('.tsv') else None
  if delimiter is None:
    # 헤더 라인에서 sniff
    first_line = text.split('\n', 1)[0]
    delimiter = '\t' if first_line.count('\t') > first_line.count(',') else ','

  reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
  if not reader.fieldnames:
    return [], ['헤더 행을 읽을 수 없습니다']

  # 원본 컬럼 → canonical 매핑
  col_map: dict[str, str] = {}
  for raw_col in reader.fieldnames:
    canon = _normalize_col(raw_col)
    if canon:
      col_map[raw_col] = canon

  if 'question' not in col_map.values() or 'answer' not in col_map.values():
    return [], [
      f'필수 컬럼 누락 — 발견된 헤더: {list(reader.fieldnames)}. '
      f'"question"/"질문" 과 "answer"/"답변" 중 어떤 이름이든 사용 가능합니다.'
    ]

  pairs: list[dict] = []
  errors: list[str] = []
  for i, row in enumerate(reader, start=2):
    mapped = {col_map[k]: v for k, v in row.items() if k in col_map}
    qa = _row_to_qa(mapped)
    if qa:
      pairs.append(qa)
    else:
      if any((row.get(k) or '').strip() for k in row):
        errors.append(f'행 {i}: question 또는 answer 가 비었거나 길이 초과')
  return pairs, errors


# ─────────────────────────────────────────────────────────
# XLSX
# ─────────────────────────────────────────────────────────

def parse_xlsx(raw: bytes, _filename: str) -> tuple[list[dict], list[str]]:
  try:
    from openpyxl import load_workbook
  except ImportError:
    return [], ['openpyxl 패키지가 설치되지 않았습니다']

  try:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
  except Exception as e:
    return [], [f'Excel 파일을 읽을 수 없습니다: {type(e).__name__}']

  ws = wb.active
  if ws is None:
    return [], ['시트가 없습니다']

  rows_iter = ws.iter_rows(values_only=True)
  try:
    header_row = next(rows_iter)
  except StopIteration:
    return [], ['빈 시트입니다']

  headers = [str(c).strip() if c is not None else '' for c in header_row]
  col_map: dict[int, str] = {}
  for idx, h in enumerate(headers):
    canon = _normalize_col(h)
    if canon:
      col_map[idx] = canon

  if 'question' not in col_map.values() or 'answer' not in col_map.values():
    return [], [
      f'필수 컬럼 누락 — 발견된 헤더: {headers}. '
      f'"question"/"질문" 과 "answer"/"답변" 중 어떤 이름이든 사용 가능합니다.'
    ]

  pairs: list[dict] = []
  errors: list[str] = []
  for i, row in enumerate(rows_iter, start=2):
    mapped: dict = {}
    for idx, canon in col_map.items():
      if idx < len(row):
        v = row[idx]
        mapped[canon] = str(v) if v is not None else ''
    qa = _row_to_qa(mapped)
    if qa:
      pairs.append(qa)
    elif any(v for v in row if v is not None and str(v).strip()):
      errors.append(f'행 {i}: question 또는 answer 가 비었거나 길이 초과')
  try:
    wb.close()
  except Exception:
    pass
  return pairs, errors


# ─────────────────────────────────────────────────────────
# JSON
# ─────────────────────────────────────────────────────────

def parse_json(raw: bytes, _filename: str) -> tuple[list[dict], list[str]]:
  text = _decode_text(raw)
  try:
    data = json.loads(text)
  except json.JSONDecodeError as e:
    return [], [f'JSON 파싱 실패: {e}']

  if isinstance(data, dict):
    data = data.get('qa_pairs') or data.get('items') or data.get('data') or []
  if not isinstance(data, list):
    return [], ['JSON 은 배열 또는 {"qa_pairs": [...]} 형태여야 합니다']

  pairs: list[dict] = []
  errors: list[str] = []
  for i, item in enumerate(data, start=1):
    if not isinstance(item, dict):
      errors.append(f'항목 {i}: 객체가 아님')
      continue
    mapped: dict = {}
    for k, v in item.items():
      canon = _normalize_col(k)
      if canon and isinstance(v, (str, int, float)):
        mapped[canon] = str(v)
    qa = _row_to_qa(mapped)
    if qa:
      pairs.append(qa)
    else:
      errors.append(f'항목 {i}: question/answer 누락 또는 길이 초과')
  return pairs, errors


# ─────────────────────────────────────────────────────────
# TXT / MD — 정규식 + LLM fallback
# ─────────────────────────────────────────────────────────

# 라인 단위 Q/A 마커 감지.
#
# 설계:
# - "Q" / "A" 는 단독 토큰이어야 함 — 뒤에 알파벳이 오면 NOT marker (예: "And", "An apple").
#   → negative lookahead `(?![A-Za-z])` 로 보장.
# - "Q1.", "Q2:" 같은 숫자 접미사 허용 (`\d{0,2}`).
# - 콜론(`:` 또는 `：`) 은 선택.
# - 마커와 본문 사이는 최소 공백 1 필요 (콜론 없는 경우).
# - 한국어/영어 긴 마커 (질문, 답변, Question, Answer 등) 도 지원.
# 번호 구분자: 콜론, 마침표, 닫는 괄호 (Q1: / Q1. / Q1) 모두 지원)
_Q_MARKER = re.compile(
  r'^\s*(?:Q(?![A-Za-z])|질문|문제|Question(?![a-z]))\d{0,2}[.:：)\]]?[ \t]+(\S.*?)\s*$',
  re.IGNORECASE,
)
_A_MARKER = re.compile(
  r'^\s*(?:A(?![A-Za-z])|답변|답(?![가-힣])|Answer(?![a-z])|응답)\d{0,2}[.:：)\]]?[ \t]+(\S.*?)\s*$',
  re.IGNORECASE,
)


def _is_meaningful_question(text: str) -> bool:
  """Q 문장이 실제로 의미 있는 질문인지 판정 (쓰레기 fragment 제거)."""
  if not text or len(text.strip()) < 5:
    return False
  # 알파벳/한글 토큰 2개 이상 필요 (빈 껍데기 "?" 이나 단어 하나짜리 거부)
  import re as _re
  tokens = _re.findall(r'[A-Za-z]{2,}|[\uac00-\ud7a3]{2,}', text)
  if len(tokens) < 2:
    return False
  # 극단적으로 짧은 영어 한 단어만 있는 경우 거부 ("self-efficacy ?" 같은 keyword fragment)
  if len(text.split()) < 3 and not any('\uac00' <= c <= '\ud7a3' for c in text):
    return False
  return True


def _normalize_text(text: str) -> str:
  """PDF 에서 추출된 텍스트의 제어문자 정규화.

  pdfplumber 는 글꼴 간 spacing 을 NULL(\\x00) 이나 다른 제어문자로 표현할 수 있다.
  이 함수는 모든 제어문자(탭/개행/CR 제외)를 공백으로 치환해 regex 매칭 가능하게 한다.
  """
  # \x09=\t, \x0a=\n, \x0d=\r 는 유지. 나머지 [\x00-\x1f] 제어문자는 공백으로.
  import re as _re
  return _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', ' ', text)


def _parse_text_regex(text: str) -> list[dict]:
  """라인 기반 Q/A 파서.

  설계:
  - Q 마커 라인을 만나면 새 쌍 시작 (이전 쌍 flush).
  - Q 마커 라인은 그 자체로 Q 본문 포함 (콜론/공백 뒤의 텍스트).
  - 다음 라인부터는 Q 본문의 continuation 으로 간주 (A 마커가 나오기 전까지).
  - A 마커가 나오면 A 본문 시작, 다음 Q 마커 또는 EOF 까지 A 본문 continuation.
  - Q 본문과 A 본문은 공백으로 join 후 정규화 (중복 공백 제거).

  보증:
  - 질문 N 과 답변 N 은 반드시 텍스트에서 물리적으로 인접한 동일 블록에서만 쌍을 이룸 (cross-block shift 절대 없음).
  - Q 또는 A 라인이 마커 없이 다음 페이지로 넘어가도 continuation 으로 합쳐짐.
  - 빈 쌍, 짧은 질문(2단어 미만), 중복 질문은 필터.
  """
  text = _normalize_text(text)
  lines = text.split('\n')
  pairs: list[dict] = []
  seen: set[str] = set()

  state: str = 'idle'  # 'idle' | 'q' | 'a'
  q_parts: list[str] = []
  a_parts: list[str] = []

  def flush():
    if q_parts and a_parts:
      q = ' '.join(' '.join(q_parts).split())
      a = ' '.join(' '.join(a_parts).split())
      if q and a and _is_meaningful_question(q) and len(q) <= 2000 and len(a) <= 5000:
        key = q.lower().strip()
        if key not in seen:
          seen.add(key)
          pairs.append({
            'question': q, 'answer': a,
            'short_answer': None, 'keywords': None, 'category': None,
          })

  for raw in lines:
    line = raw.rstrip()
    qm = _Q_MARKER.match(line)
    am = _A_MARKER.match(line)

    if qm:
      # 새 Q 시작 → 이전 쌍 flush
      flush()
      q_parts = [qm.group(1)]
      a_parts = []
      state = 'q'
      continue

    if am and state in ('q',):
      # Q → A 전이
      a_parts = [am.group(1)]
      state = 'a'
      continue

    # continuation — 현재 state 에 따라 Q 또는 A 의 다음 라인으로 누적
    stripped = line.strip()
    if not stripped:
      # 빈 줄: 단순 스킵 (블록 경계가 아님 — Q/A 마커만이 경계)
      continue
    if state == 'q':
      q_parts.append(stripped)
    elif state == 'a':
      # A 본문 continuation — 단, 다음 Q/A 마커를 잘못 먹지 않도록 위에서 이미 걸러짐
      a_parts.append(stripped)
    # state == 'idle' 이면 라인 무시 (Q 마커 없이 떠도는 텍스트)

  flush()
  return pairs


async def parse_text(raw: bytes, filename: str) -> tuple[list[dict], list[str]]:
  text = _decode_text(raw)
  if not text.strip():
    return [], ['빈 파일']

  # 1차: 정규식 (빠르고 비용 0)
  pairs = _parse_text_regex(text)
  if pairs:
    return pairs, []

  # 2차: LLM extract (자유 형식 — FAQ 페이지, 서술형 등)
  from services.llm_service import extract_qa_pairs_from_text
  extracted = await extract_qa_pairs_from_text(text, source_hint=filename)
  return extracted, [] if extracted else ['Q&A 쌍을 찾을 수 없습니다']


# ─────────────────────────────────────────────────────────
# PDF / DOCX — 표 우선, 패턴 차선, LLM 마지막
#
# 설계 원칙 (정합성 우선):
# - Q/A 는 "데이터 쌍" — 질문 N 과 답변 N 이 물리적으로 엄격히 대응되어야 한다.
# - PDF 에서 텍스트만 선형 추출하면 2단 표가 해체되어 질문/답변이 어긋난다.
# - 따라서: 1) 표 추출 → 2) 엄격 정규식 → 3) 엄격 LLM 순서.
# - LLM 은 "인접 라인에서만 쌍 만들기" 제약 + 의심 시 0 반환 원칙.
# ─────────────────────────────────────────────────────────

def _extract_pdf_tables_sync(body: bytes) -> list[list[list[str]]]:
  """PDF 에서 모든 페이지의 table 들을 추출. row[cell] 구조.
  pdfplumber 사용 — 이미 설치되어 있음 (_extract_pdf_sync 에서 사용)."""
  import io
  import pdfplumber
  all_tables: list[list[list[str]]] = []
  try:
    with pdfplumber.open(io.BytesIO(body)) as pdf:
      for page in pdf.pages:
        try:
          for tbl in page.extract_tables() or []:
            # 각 셀을 str 로, None → ''
            rows = [[(c or '').strip() for c in row] for row in tbl if row]
            # 완전히 빈 행 제거
            rows = [r for r in rows if any(c for c in r)]
            if rows:
              all_tables.append(rows)
        except Exception:
          continue
  except Exception:
    pass
  return all_tables


def _extract_docx_tables_sync(body: bytes) -> list[list[list[str]]]:
  import io
  try:
    from docx import Document
  except ImportError:
    return []
  try:
    doc = Document(io.BytesIO(body))
  except Exception:
    return []
  all_tables: list[list[list[str]]] = []
  for tbl in doc.tables:
    rows: list[list[str]] = []
    for tr in tbl.rows:
      row = [(c.text or '').strip() for c in tr.cells]
      if any(row):
        rows.append(row)
    if rows:
      all_tables.append(rows)
  return all_tables


def _table_to_qa_pairs(table: list[list[str]]) -> list[dict]:
  """2+ 열 표를 Q/A 쌍으로 변환.
  1) 첫 행이 헤더면 헤더 이름으로 컬럼 매핑 (question/질문/Q, answer/답변/A, ...)
  2) 헤더가 없으면 첫 두 열을 question/answer 로 가정.
  3) 각 행마다 엄격히 같은 row 의 셀만 쌍으로 묶음 (cross-row 금지).
  """
  if not table or len(table) == 0:
    return []
  n_cols = max(len(r) for r in table)
  if n_cols < 2:
    return []

  # 헤더 감지 시도
  header = table[0]
  col_map: dict[str, int] = {}
  for idx, cell in enumerate(header):
    canon = _normalize_col(cell)
    if canon:
      col_map[canon] = idx
  has_header = 'question' in col_map and 'answer' in col_map
  if has_header:
    data_rows = table[1:]
    q_idx = col_map['question']
    a_idx = col_map['answer']
    sa_idx = col_map.get('short_answer')
    kw_idx = col_map.get('keywords')
    cat_idx = col_map.get('category')
  else:
    # 헤더 없음 → 첫 두 열을 Q/A 로 가정
    data_rows = table
    q_idx = 0
    a_idx = 1
    sa_idx = kw_idx = cat_idx = None

  pairs: list[dict] = []
  for row in data_rows:
    if len(row) <= max(q_idx, a_idx):
      continue
    q = (row[q_idx] or '').strip()
    a = (row[a_idx] or '').strip()
    if not q or not a:
      continue
    # 한글 Q/A 의 경우 셀 안에 \n 이 섞여 있을 수 있음 — 공백으로 정규화
    q = ' '.join(q.split())
    a = ' '.join(a.split())
    if len(q) > 2000 or len(a) > 5000:
      continue
    pairs.append({
      'question': q,
      'answer': a,
      'short_answer': (row[sa_idx].strip() or None) if sa_idx is not None and len(row) > sa_idx else None,
      'keywords': (row[kw_idx].strip() or None) if kw_idx is not None and len(row) > kw_idx else None,
      'category': (row[cat_idx].strip() or None) if cat_idx is not None and len(row) > cat_idx else None,
    })
  return pairs


async def parse_document(raw: bytes, filename: str) -> tuple[list[dict], list[str]]:
  """문서(PDF/DOCX)에서 Q&A 추출.
  우선순위: 1) 표 구조 → 2) 정규식 패턴 → 3) LLM 엄격 추출.
  각 단계는 "물리적 인접성" 보장 — 질문 N 과 답변 N 이 같은 표 row 또는 인접 라인에서만 쌍을 이룸.
  """
  import asyncio as _asyncio
  ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

  # ── Step 1: 표 구조 추출 (가장 정확) ──
  try:
    if ext == 'pdf':
      tables = await _asyncio.to_thread(_extract_pdf_tables_sync, raw)
    elif ext in ('docx', 'doc'):
      tables = await _asyncio.to_thread(_extract_docx_tables_sync, raw)
    else:
      tables = []
  except Exception as e:
    logger.warning(f'table extraction failed for {filename}: {e}')
    tables = []

  table_pairs: list[dict] = []
  for tbl in tables:
    table_pairs.extend(_table_to_qa_pairs(tbl))

  if len(table_pairs) >= 2:
    # 2개 이상 표에서 잡히면 표 기반이 가장 신뢰할 수 있음 → 이것만 사용
    logger.info(f'{filename}: table-based extraction → {len(table_pairs)} pairs')
    return table_pairs, []

  # ── Step 2: 선형 텍스트 추출 + 정규식 ──
  from services.extractors import extract, ExtractError
  try:
    result = await extract(raw, filename_hint=filename)
  except ExtractError as e:
    if table_pairs:
      return table_pairs, []
    return [], [f'문서 텍스트 추출 실패: {e}']

  if not result.text or not result.text.strip():
    if table_pairs:
      return table_pairs, []
    return [], ['문서에서 텍스트를 추출할 수 없습니다']

  regex_pairs = _parse_text_regex(result.text)
  if len(regex_pairs) >= 3:
    logger.info(f'{filename}: regex-based extraction → {len(regex_pairs)} pairs')
    return regex_pairs, []

  # ── Step 3: LLM 엄격 추출 ──
  from services.llm_service import extract_qa_pairs_from_text
  extracted = await extract_qa_pairs_from_text(result.text, source_hint=filename)
  if extracted:
    logger.info(f'{filename}: LLM-based extraction → {len(extracted)} pairs')
    return extracted, []

  if table_pairs or regex_pairs:
    return table_pairs or regex_pairs, []
  return [], ['Q&A 쌍을 찾을 수 없습니다. 표 구조 / Q:A: 패턴 / 자유 형식 모두 실패.']


# ─────────────────────────────────────────────────────────
# Public dispatcher
# ─────────────────────────────────────────────────────────

async def parse_qa_file(raw: bytes, filename: str) -> tuple[list[dict], list[str]]:
  """
  파일 내용에서 Q&A 쌍을 추출.

  Returns: (pairs, errors)
    pairs: list[{question, answer, short_answer, keywords, category}]
    errors: 사용자 노출 가능한 경고/에러 메시지
  """
  if not filename:
    return [], ['파일명이 없습니다']
  ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
  kind = PARSER_MAP.get(ext)
  if not kind:
    return [], [f'지원하지 않는 형식: .{ext}. 지원 포맷: {", ".join(sorted(PARSER_MAP.keys()))}']

  try:
    if kind == 'tabular':
      return parse_tabular(raw, filename)
    if kind == 'xlsx':
      return parse_xlsx(raw, filename)
    if kind == 'json':
      return parse_json(raw, filename)
    if kind == 'text':
      return await parse_text(raw, filename)
    if kind == 'document':
      return await parse_document(raw, filename)
  except Exception as e:
    logger.exception(f'parse_qa_file failed: {filename}')
    return [], [f'파일 파싱 실패: {type(e).__name__}: {e}']

  return [], [f'지원하지 않는 형식: .{ext}']
